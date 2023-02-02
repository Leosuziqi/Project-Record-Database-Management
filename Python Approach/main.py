import win32com.client, datetime
from datetime import date
from dateutil.parser import *
import calendar
import pandas as pd
# Data fetching from Excel files
import pathlib
# Author: Leo Su
# Data 19-Dec-2022

from tkinter import *
import tkinter as tk
import numpy as np
import openpyxl,xlrd
from tkinter import Tk, mainloop,TOP
from openpyxl import Workbook
#import cv2
from tkinter import ttk
from tkinter.messagebox import showinfo
from fetch_calendar import *
import os, os.path
import warnings

start_date_column = 48
end_date_column = 49
start_flag=0
if(start_flag==0):
    xl = win32com.client.DispatchEx("Excel.Application")
    wb = xl.workbooks.open("C:\\Users\\leo.su\\PycharmProjects\\outlook\\demo\\New Project.xlsm")
    xl.Visible = True
    xl.run("'New Project.xlsm'!update_from_master")

    wb.SaveAs(Filename="C:\\Users\\leo.su\\PycharmProjects\\outlook\\demo\\temp.xlsm")
    wb.Close()
    xl.Quit()
    print("VBA Done.")
    #get_outlood_data
    outlook_data = fetch_calendar()
    #print(outlook_data)
    #output=fetch_calendar()
    print("Calendar fetched.")
    #Load file
    excel_path = ".\\demo\\temp.xlsm"
    #excel_path = r"I:\Support Group\Service\Freezer Job Record.xlsx"
    file = openpyxl.load_workbook(excel_path,data_only=True, keep_vba=True)
    sheet = file['New Projects']
    #print(2)
    itr_row = 2

    for cell_2 in sheet.iter_rows(min_row=2, max_row=1000, min_col=1,max_col=56, values_only=FALSE):
        for index, row in outlook_data.iterrows():
            #print(itr_row)
            if str(row['project_id']) == str(sheet[itr_row][3].value):
                print(row['project_id'])
                #s=sheet.cell(row=itr_row, column=start_date_column + 1).value
                sheet.cell(row=itr_row, column=start_date_column).value = (str(row['Start_Date']))
                sheet.cell(row=itr_row, column=end_date_column).value = row['End_Date']
            break
        itr_row=itr_row+1

    warnings.simplefilter("ignore")
    file.save(".\\demo\\modified New Project.xlsm")
    start_flag=1



main=Tk()
main.title('Search_Page')
main.geometry('800x500')
main.config(highlightbackground="black", highlightthickness=2)

"""
file = pathlib.Path("test.xlsx")
if file.exists():
    pass
else:
    #Erro Message: File not Existed.
    error_page=Tk()
    error_page.title('error')
    error_page.geometry("200*100")
    error_label = tk.Label(error_page, text="File not Existed.")
    error_label.place(anchor="center")
"""
excel_path = r".\Freezer Job Record.xlsx"
#excel_path = r"I:\Support Group\Service\Freezer Job Record.xlsx"

#frame1 = LabelFrame(main,text = 'Search information:')
#frame1.pack(expand = 'yes', fill = 'both')

output_table = fetch_calendar()

item1_name=Label(main,text="Project #:")
item1_name.place(x=50,y=70)
item2_name=Label(main,text="Company Name:")
item2_name.place(x=50,y=100)
item3_name=Label(main,text="Location:")
item3_name.place(x=50,y=130)

ProjectNum=Entry(main)
ProjectNum.place(x=250,y=70)
CompanyName=Entry(main)
CompanyName.place(x=250,y=100)
Location=Entry(main)
Location.place(x=250,y=130)

Button(main,text="Search",command=search).place(x=230, y=400)
# Button for closing
Button(main, text="Exit", command=main.destroy).place(x=460, y=400)


main.mainloop()
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
