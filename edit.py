# Data fetching from Excel files
import pathlib
# Author: Leo Su
# Data 19-Dec-2022

from tkinter import *
import tkinter as tk
import numpy as np
import openpyxl,xlrd
from tkinter import Tk, mainloop,TOP
from openpyxl import Workbook
#import cv2
from tkinter import ttk
from tkinter.messagebox import showinfo

main=Tk()
main.title('Search_Page')
main.geometry('800x350')
main.config(highlightbackground="black", highlightthickness=2)

global row_to_change
global col_to_change
global value_to_change

#Load file
excel_path = r".\Freezer Job Record.xlsx"
#excel_path = r"I:\Support Group\Service\Freezer Job Record.xlsx"

#File not Existed error
if file.exists():
    pass
else:
    #Erro Message: File not Existed.
    error_page=Tk()
    error_page.title('error')
    error_page.geometry("200*100")
    error_label = tk.Label(error_page, text="File not Existed.")
    error_label.place(anchor="center")
    
file = openpyxl.load_workbook(excel_path, data_only=True)
sheet = file['Freezer Job Record']

#Search Function
def search():
    #search
    rows = []
    id_flag = FALSE
    Company_flag = FALSE
    Location_flag = FALSE
    if ProjectNum.get():
        search_id=ProjectNum.get()
        id_flag = TRUE

    if CompanyName.get():
        search_Company=CompanyName.get()
        Company_flag = TRUE

    if Location.get():
        search_Location=Location.get()
        Location_flag = TRUE

    ProjectNum.configure(state=tk.NORMAL)
    CompanyName.configure(state=tk.NORMAL)
    Location.configure(state=tk.NORMAL)
    #email.configure(state=tk.NORMAL)


    #Save every matched row into cols, then append all cols to rows[] as the searched result
    cols=[]
    for index, cell in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=22,values_only=FALSE):    #Iteration in rows
        #values_only?
        if not (id_flag|Company_flag|Location_flag):
            continue
        #check if one or two or three inserted data matched
        if (id_flag and (str(search_id).lower() in str(sheet[itr][0].value).lower())) or id_flag == FALSE:
            #print(id_flag)
            #print(sheet[itr][0].value)
            if (Company_flag and (str(search_Company).lower() in str(sheet[itr][1].value).lower())) or Company_flag == FALSE:
                if (Location_flag and str(search_Location).lower() in str(sheet[itr][2].value).lower()) or Location_flag == FALSE:
                    for index in cell:
                        #print(index)
                        cols.append(index.value)
                    rows.append(cols)
                    cols=[]

    #print(len(rows))

    #display the homepage
    main.withdraw()
    headers = []
    changed_rows = []
    for header in sheet[1]:
        if len(headers)<22:
            headers.append(header)

    result = tk.Tk()
    result.title('Search Result')
    result.geometry('1100x400')

    # Setup of Scrollbars in X and Y directions
    scrollbary = ttk.Scrollbar(result, orient=tk.VERTICAL)
    scrollbarx = ttk.Scrollbar(result, orient=tk.HORIZONTAL)
    treeview = ttk.Treeview(result, columns=headers, show='headings')
    treeview.place(x=40,y=40,width=1000,height=250)
    treeview.configure(yscrollcommand=scrollbary.set, xscrollcommand = scrollbarx.set)
    result.update()
    scrollbary.configure(command=treeview.yview)
    scrollbarx.configure(command=treeview.xview)

    scrollbary.place(x=740,y=40,width=20,height=250)
    scrollbarx.place(x=40,y=290,width=1000,height=20)

    # Display data in treeview
    treeview.configure(columns=headers)
    for index in headers:
        treeview.heading(index, text=index.value,anchor=W)
        treeview.column(index,stretch=NO,width=80)
    for content in rows:
        treeview.insert('', tk.END, values=content)

    """
    #columns = ('Project #', 'Company Name', 'Location')

    # Initalize a Treeview

    treeview = ttk.Treeview(result_frame, columns=headers, show='headings')
    for index in headers:
        treeview.heading(index, text=index.value)

    for content in rows:
        treeview.insert('', tk.END, values=content)

    # add scrollbars
    scrollbary = ttk.Scrollbar(result_frame, orient=tk.VERTICAL)
    scrollbarx = ttk.Scrollbar(result_frame, orient=tk.HORIZONTAL)

    scrollbary.config(command=treeview.yview)
    scrollbarx.config(command=treeview.xview)

    result_frame.grid(column=0, row=0, sticky=(N, E))
    scrollbarx.grid(row=1, column=0, sticky=tk.E+tk.W)
    treeview.grid(row=0, column=0, sticky=tk.N + tk.W)
    scrollbary.grid(row=0, column=1, sticky=tk.N + tk.S)


    treeview.configure(yscrollcommand=scrollbary.set, xscrollcommand = scrollbarx.set)
    """
    #command for back_to_search button
    def back_to_search():
        main.deiconify()
        result.destroy()

    # command for search_close button
    def search_close():
        main.destroy()
        result.destroy()
        
    #command for back_to_search button
    def back_to_search():
        main.deiconify()
        result.destroy()

    # command for search_close button
    def search_close():
        main.destroy()
        result.destroy()
        
    def edit():
        # Get selected item to Edit
        for row_changes in changed_rows:
            for ,index, cell_2 in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=22, values_only=FALSE):
                if str(row_changes[0]) in str(sheet[itr_row][0].value):
                    sheet.cell(row=itr_row,column=row_changes[1]+1).value = row_changes[2]
                    break

        update_file()
        result.destroy()
        search()

    def selectItemtoEdit(event):
        global row_to_change
        global col_to_change
        global value_to_change
        region = treeview.identify_region(event.x,event.y)
        if region not in ("tree","cell"):
            return

        #Record Clicked cell location
        column = treeview.identify_column(event.x)
        column_index=int(column[1:])-1
        iid=treeview.focus()
        iid_index = int(iid[1:])-1
        selected_value=treeview.item(iid)
        selected_text=selected_value.get("values")[column_index]

        #print(selected_text)

        column_box=treeview.bbox(iid,column)
        entry_edit= tk.Entry(result)

        entry_edit.editing_column_index=column_index
        entry_edit.editing_row_index=iid_index

        """
        def on_enter(self):
            global value_to_change
            value_to_change=entry_edit.get()

            entry_edit.destroy()
        """

        #When click outside of edited cell
        def on_focus_out(self):
            # change table view
            if entry_edit.winfo_exists():
                new_text = entry_edit.get()

                selected_iid = entry_edit.editing_row_index
                selected_column = entry_edit.editing_column_index

                #print(1)
                #print(selected_iid)
                #print(selected_column)
                if selected_column == -1:
                    treeview.item(selected_iid,text=new_text)
                else:
                    current_value = treeview.item(treeview.selection())['values']

                    if current_value[selected_column] != new_text:
                        current_value[selected_column] = new_text
                        treeview.item(treeview.selection(),values=current_value)

                        # save to the list of changes
                        #chnages in form of [project_ID,column_to_change,value_to_change]
                        changes = [current_value[0], selected_column, new_text]
                        changed_rows.append(changes)
                        print(changed_rows)

                entry_edit.destroy()

        entry_edit.insert(0,selected_text)
        #Press "Enter" to save changes
        entry_edit.bind("<Return>",on_focus_out)

        entry_edit.place(x=column_box[0]+40,
                         y=column_box[1]+40,
                         w=column_box[2],
                         h=column_box[3])
        row_to_change = iid_index
        col_to_change = column_index

        #Click other area to save changes
        treeview.bind("<Button-1>", on_focus_out)
        scrollbary.bind("<Button-1>", on_focus_out)
        scrollbarx.bind("<Button-1>", on_focus_out)


    def edit():
        # Get selected item to Edit
        for row_changes in changed_rows:
            for index, cell_2 in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=22, values_only=FALSE):
                if str(row_changes[0]) in str(sheet[itr_row][0].value):
                    sheet.cell(row=itr_row,column=row_changes[1]+1).value = row_changes[2]
                    break
                    
        update_file()
        result.destroy()
        search()

    def delete():
        # Get selected item to Delete
        update_file()

    def update_file():
        #Get selected item to Delete
        #save to .xlsx
        file.save(excel_path)

    re_search=ttk.Button(result, text="Return to Search", command=back_to_search)
    re_search.place(relx=0.6,rely=0.855,width=100,height=40)
    exit_2 = ttk.Button(result, text="Close", command=search_close)
    exit_2.place(relx=0.8, rely=0.855, width=100, height=40)


    # Add Buttons to Edit and Delete the Treeview items
    treeview.bind('<Double-1>', selectItemtoEdit)
    edit_btn = ttk.Button(result, text="Save to Excel File", command=edit)
    edit_btn.place(relx=0.2,rely=0.855,width=100,height=40)
    #del_btn = ttk.Button(result, text="Delete", command=delete)
    #del_btn.place(relx=0.4,rely=0.855,width=100,height=40)
#frame1 = LabelFrame(main,text = 'Search information:')
#frame1.pack(expand = 'yes', fill = 'both')

#Labels and Entries in Home page
item1_name=Label(main,text="Project #:")
item1_name.place(x=50,y=70)
item2_name=Label(main,text="Company Name:")
item2_name.place(x=50,y=100)
item3_name=Label(main,text="Location:")
item3_name.place(x=50,y=130)
item4_name=Label(main,text="*Please Enter AT LEAST One Element to Search")
item4_name.place(x=250,y=40)

ProjectNum=Entry(main, width=70)
ProjectNum.place(x=250,y=70)
CompanyName=Entry(main, width=70)
CompanyName.place(x=250,y=100)
Location=Entry(main, width=70)
Location.place(x=250,y=130)


#Buttons in Home page
Button(main,text="Search",command=search,font=15).place(x=230, y=250)
# Button for closing
Button(main, text="Exit", command=main.destroy,font=15).place(x=460, y=250)

main.mainloop()
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
