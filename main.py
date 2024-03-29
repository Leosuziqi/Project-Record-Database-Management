# Author: Leo Su
# Data 19-Dec-2022

import numpy as np
import openpyxl
from tkinter import Tk, mainloop,TOP, ttk
from openpyxl import Workbook
from tkinter.messagebox import showinfo
from fetch_calendar import *
import os, os.path
import warnings
import win32com.client, datetime
from datetime import date
from dateutil.parser import *
import calendar
import pandas as pd
import pathlib

from fetch_calendar import *
from search import *
from edit import *
from TreeviewAction import *

## Run Excel Macro
xl = win32com.client.DispatchEx("Excel.Application")
wb = xl.workbooks.open("C:\\Users\\leo.su\\PycharmProjects\\outlook\\demo\\New Project.xlsm")
xl.Visible = True
xl.run("'New Project.xlsm'!update_from_master")

wb.SaveAs(Filename="C:\\Users\\leo.su\\PycharmProjects\\outlook\\demo\\temp.xlsm")
wb.Close()
xl.Quit()
print("VBA Done.")


## get outlook data
outlook_data = fetch_calendar()
#print(outlook_data)
#output=fetch_calendar()
print("Calendar fetched.")

## Uplaad changes
#Load file
excel_path = ".\\demo\\temp.xlsm"
#excel_path = r"I:\Support Group\Service\Freezer Job Record.xlsx"
file = openpyxl.load_workbook(excel_path,data_only=True, keep_vba=True)
sheet = file['New Projects']

# fetch_calendar returns list of rows in the format of:
# ("project_id", "Duration",  "Subject", "Start_Date","Organizer", "End_Date")
outlook_data = fetch_calendar()

#iterate through excel sheet and find same project_ID from outlook data
for rows_in_sheet in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1,max_col=56, values_only=FALSE):
    for index, row in outlook_data.iterrows():
        #print(index)
        if str(row['project_id']) == str(sheet[itr_row][3].value):
            #print(row['project_id'])
            # change start_date and end_date
            sheet.cell(row=itr_row, column=start_date_column).value = (str(row['Start_Date']))
            sheet.cell(row=itr_row, column=end_date_column).value = row['End_Date']
        break

# save file and ignore warnings
warnings.simplefilter("ignore")
file.save(".\\demo\\modified New Project.xlsm")


## Display UI
main=Tk()
main.title('Search_Page')
main.geometry('800x500')
main.config(highlightbackground="black", highlightthickness=2)

excel_path = r".\Freezer Job Record.xlsx"
#excel_path = r"I:\Support Group\Service\Freezer Job Record.xlsx"

#frame1 = LabelFrame(main,text = 'Search information:')
#frame1.pack(expand = 'yes', fill = 'both')

# design label and entry
item1_name=Label(main,text="Project #:")
item1_name.place(x=50,y=70)
item2_name=Label(main,text="Company Name:")
item2_name.place(x=50,y=100)
item3_name=Label(main,text="Location:")
item3_name.place(x=50,y=130)

ProjectNum=Entry(main)
ProjectNum.place(x=250,y=70)
CompanyName=Entry(main)
CompanyName.place(x=250,y=100)
Location=Entry(main)
Location.place(x=250,y=130)

Button(main,text="Search",command=search).place(x=230, y=400)
# Button for closing
Button(main, text="Exit", command=main.destroy).place(x=460, y=400)


main.mainloop()
